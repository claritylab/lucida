

from openpyxl import load_workbook
import ezodf

def modif_xls(server_cost, server_power):
#  wb = load_workbook('TCO_siri.xlsx',data_only=True)
  wb = load_workbook('TCO_siri.xlsx')
  #print wb.get_sheet_names()
  ws_in = wb.get_sheet_by_name('TCO Inputs')
#  ws_out = wb.get_sheet_by_name('TCO Outputs')

  # set values for each platform
  ws_in['B15'] = float(server_cost)
  ws_in['B18'] = float(server_power)

  wb.save('TCO_siri.xlsx')

def read_xls():
#  wb = load_workbook('TCO_siri.xlsx')
  wb = load_workbook('TCO_siri.xlsx',data_only=True)
  ws_in = wb.get_sheet_by_name('TCO Inputs')
  ws_out = wb.get_sheet_by_name('TCO Outputs')
 
  # TCO, num of servers, power_budget
  return ws_out['B14'].value, ws_in['E10'].value, ws_in['B10'].value

def modif_ods(server_cost, server_power):
  spreadsheet = ezodf.opendoc("TCO_siri.ods")
  ws_in = spreadsheet.sheets['TCO Inputs']
  ws_in['B15'].set_value(server_cost)
  ws_in['B18'].set_value(server_power)
  spreadsheet.save()

def read_ods():
  spreadsheet = ezodf.opendoc("TCO_siri.ods")
  ws_in = spreadsheet.sheets['TCO Inputs']
  ws_out = spreadsheet.sheets['TCO Outputs']
 
  # TCO, num of servers, power_budget
  return ws_out['B14'].value, ws_in['E10'].value, ws_in['B10'].value


# either gmm/hmm or dnn/hmm as voice front-end
# img/match and img/ocr as image front-end
# openephyra as question-answering backend

components = ['gmm/hmm', 'dnn/hmm', 'openephyra', 'img/match', 'img/ocr']

query_dist_ = [0.0, 0.0, 0.0, 0.0, 0.0]

kernels = ['']

CPU_server_cost = 798.00
CPU_server_power = 300.0

GPU_server_cost = 1159.99
GPU_server_power = 500.0

# plot different power budgets?
#print 'power budget', ws_in['B10'].value

modif_xls(CPU_server_cost, CPU_server_power)
#modif_xls(GPU_server_cost, GPU_server_power)

#modif_ods(CPU_server_cost, CPU_server_power)
#modif_ods(GPU_server_cost, GPU_server_power)

tco, num_servers, power_budget = read_xls()
#tco, num_servers, power_budget = read_ods()

print 'CPU only configuration'
print 'TCO:', tco
print 'Num servers:', num_servers
print 'Power budget:', power_budget




